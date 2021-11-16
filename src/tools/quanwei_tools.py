# coding=utf-8
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from copy import copy
import tkinter.filedialog
import warnings
import datetime
import os

warnings.filterwarnings('ignore')
window = Tk()
# 设置窗口大小
winWidth = 800
winHeight = 600
# 获取屏幕分辨率
screenWidth = window.winfo_screenwidth()
screenHeight = window.winfo_screenheight()
x = int((screenWidth - winWidth) / 2)
y = int((screenHeight - winHeight) / 2)
# 设置主窗口标题
window.title('全委表单差异处理工具')
# 设置窗口初始位置在屏幕居中
window.geometry('%sx%s+%s+%s' % (winWidth, winHeight, x, y))
# 设置窗口宽高固定
window.resizable(0, 0)
# 省会和城市所属的事业部
provinceCityDict = {
    '浙江-嘉兴': '事业一部',
    '浙江-湖州': '事业一部',
    '浙江-金华': '事业一部',
    '浙江-丽水': '事业一部',
    '浙江-衢州': '事业一部',
    '安徽': '事业一部',
    '湖北': '事业一部',
    '广东': '事业一部',
    '广西': '事业一部',
    '海南': '事业一部',
    '福建': '事业一部',
    '浙江-温州': '事业二部',
    '浙江-台州': '事业二部',
    '山东': '事业二部',
    '重庆': '事业二部',
    '四川': '事业二部',
    '贵州': '事业二部',
    '云南': '事业二部',
    '浙江-绍兴': '事业三部',
    '浙江-宁波': '事业三部',
    '浙江-舟山': '事业三部',
    '江苏': '事业三部',
    '河北': '事业三部',
    '北京': '事业三部',
    '黑龙江': '事业三部',
    '吉林': '事业三部',
    '辽宁': '事业三部',
    '天津': '事业三部',
    '陕西': '事业三部',
    '山西': '事业三部',
    '新疆': '事业三部',
    '青海': '事业三部',
    '甘肃': '事业三部',
    '宁夏': '事业三部',
    '河南': '事业三部',
    '内蒙古': '事业三部',
    '浙江-杭州': '事业四部',
    '上海': '事业四部',
    '湖南': '事业四部',
    '江西': '事业四部',
}


# 更新窗口上的文本显示内容
def update_text(content):
    """
    更新窗口上的文本显示内容
    :param content: 文本字符串
    :return:
    """
    text.insert(INSERT, content + '\n')
    window.update()


# 清除窗口上的文本显示内容
def clear_text():
    """
    清除窗口上的文本显示内容
    :return:
    """
    text.delete('1.0', 'end')
    window.update()


# 获取工作簿从某一列开始的多行的内容
def get_row_value(ws, row=1, col=1):
    """
    获取工作簿从某一列开始的多行的内容
    :param ws: 工作簿描述符（标识）
    :param row: 行号
    :param col: 列号
    :return: 多行的内容
    """
    rowList = []
    for i in range(col, ws.max_column + 1):
        val = ws.cell(row=row, column=i).value
        rowList.append(val)
    return rowList


# 获取工作簿从某一行开始的多列的内容
def get_col_value(ws, row=1, col=1):
    """
    获取工作簿从某一行开始的多列的内容
    :param ws: 工作簿描述符（标识）
    :param row: 行号
    :param col: 列号
    :return: 多列的内容
    """
    colList = []
    for i in range(row, ws.max_row + 1):
        val = ws.cell(row=i, column=col).value
        colList.append(val)
    return colList


# 复制内容
def copy_content(source_ws, target_ws, row_index, row, target_row_index):
    """
    获取工作簿从某一行开始的多列的内容
    :param source_ws: 源文件工作簿描述符（标识）
    :param target_ws: 目标文件工作簿描述符（标识）
    :param row_index: 源文件行下标（从0开始）
    :param row: 源文件行信息
    :param target_row_index: 目标文件行下标（从0开始）
    :return:
    """
    for col_index, cell in enumerate(row):
        # 复制网格宽度
        target_ws.column_dimensions[get_column_letter(col_index + 1)].width = source_ws.column_dimensions[
            get_column_letter(col_index + 1)].width
        # 复制网格高度
        target_ws.row_dimensions[target_row_index + 1].height = source_ws.row_dimensions[row_index + 1].height
        # 复制网格内容
        target_ws.cell(row=target_row_index + 1, column=col_index + 1, value=cell.value)
        # 设置单元格格式
        source_cell = source_ws.cell(row_index + 1, col_index + 1)
        target_cell = target_ws.cell(target_row_index + 1, col_index + 1)
        target_cell.fill = copy(source_cell.fill)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)


# 在最后一列新增一列内容
def insert_col(ws, row, col, content, col_width=15, font_name='宋体', font_size=10, font_color='00000000', is_bold=False,
               is_italic=False, fg_color='00d9d2e9', bg_color='00d9d2e9', side_color='00000000'):
    """
    获取工作簿从某一行开始的多列的内容
    :param ws: 文件工作簿描述符（标识）
    :param row: 文件行号
    :param col: 文件列号
    :param content: 单元格内容
    :param col_width: 列宽
    :param font_name: 字体
    :param font_size: 字体大小
    :param font_color: 字体颜色
    :param is_bold: 是否粗体
    :param is_italic: 是否斜体
    :param fg_color: 前景色
    :param bg_color: 后景色
    :param side_color: 边框颜色
    :return:
    """
    cell = ws.cell(row, col)
    # 设置单元格内容
    cell.value = content
    # 设置单元格宽度
    ws.column_dimensions[get_column_letter(col)].width = col_width
    # 设置单元格字体
    cell.font = Font(name=font_name, size=font_size, color=Color(rgb=font_color), b=is_bold, i=is_italic)
    # 设置单元格对齐（水平居中，垂直居中）
    cell.alignment = Alignment(horizontal='center', vertical='center')
    # 设置边框
    side = Side(style='thin', color=Color(rgb=side_color))
    cell.border = Border(left=side, right=side, top=side, bottom=side)
    # 设置单元格填充
    cell.fill = PatternFill(patternType='lightDown', fgColor=Color(rgb=fg_color), bgColor=Color(rgb=bg_color))


# 比对文件差异
def compare_file_diff(wb1, wb2, sheet_name, project_name, province_name, city_name, file_name1, file_name2):
    """
    获取工作簿从某一行开始的多列的内容
    :param wb1: 文件1工作簿描述符（标识）
    :param wb2: 文件2工作簿描述符（标识）
    :param sheet_name: 工作簿名称
    :param project_name: 项目标题名称
    :param province_name: 省会标题名称
    :param city_name: 城市标题名称
    :param file_name1: 文件1全路径
    :param file_name2: 文件2全路径
    :return:
    """
    ws1 = wb1[sheet_name]
    ws2 = wb2[sheet_name]
    rowList1 = get_row_value(ws1, row=1)
    rowList2 = get_row_value(ws2, row=1)
    try:
        index1 = rowList1.index(project_name)
        index2 = rowList2.index(project_name)
        province_index = rowList2.index(province_name)
        city_index = rowList2.index(city_name)
    except ValueError:
        errStr = '文件：' + file_name1 + ' 没有指定的列名：' + project_name + '\n' + \
                 '文件：' + file_name2 + ' 没有指定的列名：' + project_name + '\n'
        messagebox.showerror('错误', errStr)
    else:
        colList1 = get_col_value(ws1, row=2, col=index1 + 1)
        colList2 = get_col_value(ws2, row=2, col=index2 + 1)
        # 求差集
        colList = list(set(colList1).difference(set(colList2)))
        # 新建工作表
        wb = Workbook()
        wb[wb.sheetnames[0]].title = sheet_name
        ws = wb[sheet_name]
        # ws = wb.create_sheet(sheet_name)
        # 往新建工作表中复制差集内容
        for rowIndex, row in enumerate(ws1.iter_rows()):
            if rowIndex == 0:
                copy_content(ws1, ws, rowIndex, row, 0)
            else:
                cellValue = ws1.cell(row=rowIndex + 1, column=row[index1].column).value
                if cellValue in colList:
                    copy_content(ws1, ws, rowIndex, row, ws.max_row)

        for rowIndex, row in enumerate(ws2.iter_rows()):
            cellValue = ws2.cell(row=rowIndex + 1, column=row[index2].column).value
            if cellValue in colList:
                copy_content(ws2, ws, rowIndex, row, ws.max_row)
        # 处理省会和城市
        target_column = ws.max_column + 1
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                insert_col(ws=ws, row=1, col=target_column, content='所属事业部', is_bold=True)
            else:
                provinceVal = ws.cell(row=i + 1, column=row[province_index].column).value
                cityVal = ws.cell(row=i + 1, column=row[city_index].column).value
                if provinceVal == '浙江':
                    key = provinceVal + '-' + cityVal
                    if key in provinceCityDict:
                        insert_col(ws=ws, row=i + 1, col=target_column, content=provinceCityDict[key])
                    else:
                        insert_col(ws=ws, row=i + 1, col=target_column, content='找不到部门数据，请检查配置')
                else:
                    if provinceVal in provinceCityDict:
                        insert_col(ws=ws, row=i + 1, col=target_column, content=provinceCityDict[provinceVal])
                    else:
                        insert_col(ws=ws, row=i + 1, col=target_column, content='找不到部门数据，请检查配置')

        # 保存新建工作表
        path, fileName = os.path.split(file_name1)
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d')
        if '物业全委' in fileName:
            newFileName = os.path.join(path, '物业全委' + nowTime + '.xlsx')
        elif '物业锁定' in fileName:
            newFileName = os.path.join(path, '物业锁定' + nowTime + '.xlsx')
        else:
            newFileName = os.path.join(path, '新建工作表' + nowTime + '.xlsx')
        wb.save(newFileName)
        wb1.close()
        wb2.close()
        messagebox.showinfo('提示', '文件处理完成！！！')


def handel_files():
    clear_text()
    fileNames = tkinter.filedialog.askopenfilenames(filetypes=[('xlsx files', '*.xlsx')])
    if len(fileNames) == 2 and ((('物业全委' in fileNames[0]) and ('物业全委' in fileNames[1])) or (
            ('物业锁定' in fileNames[0]) and ('物业锁定' in fileNames[1]))):
        # 展示已选文件
        update_text('已选文件：' + str(fileNames[0]))
        update_text('已选文件：' + str(fileNames[1]))
        # 打开xlsx表
        wb1 = load_workbook(fileNames[0])
        wb2 = load_workbook(fileNames[1])
        # 已存在的全部工作簿
        sheetNames1 = wb1.sheetnames
        sheetNames2 = wb2.sheetnames
        # 处理指定的工作簿
        if ('新签表汇总' in sheetNames1) and ('新签表汇总' in sheetNames2):
            compare_file_diff(wb1, wb2, '新签表汇总', '项目名称', '项目所在省', '项目所在市', fileNames[0], fileNames[1])
        elif ('预估签约表' in sheetNames1) and ('预估签约表' in sheetNames2):
            compare_file_diff(wb1, wb2, '预估签约表', '项目名称', '项目所在省', '项目所在市', fileNames[0], fileNames[1])
        else:
            errStr = '文件：' + str(fileNames[0]) + ' 没有需要处理的工作簿或工作簿名称不正确' + '\n' + \
                     '文件：' + str(fileNames[1]) + ' 没有需要处理的工作簿或工作簿名称不正确' + '\n'
            messagebox.showerror('错误', errStr)
    elif len(fileNames) == 0:
        messagebox.showinfo('提示', '未选择任何文件！！！')
    else:
        errStr = '只允许同时选择2个同类型文件\n已选文件：'
        for fileName in fileNames:
            errStr = errStr + str(fileName) + '\n'
        messagebox.showerror('错误', errStr)


tkinter.Button(window, text='选择文件', command=handel_files).pack()
text = tkinter.Text(window)
text.insert(INSERT, '')
text.pack()
window.mainloop()
