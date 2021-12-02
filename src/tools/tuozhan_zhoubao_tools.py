# coding=utf-8
import datetime
from tkinter import messagebox, DISABLED, NORMAL
from openpyxl import Workbook, load_workbook
import tkinter.filedialog
import os

from openpyxl.worksheet.worksheet import Worksheet

from StudyPython.src import tkinter_utils, datetime_utils, xlsx_utils

window = tkinter_utils.createWindow('拓展信息周报工具')
# 省会和城市所属的事业部
provinceCityDict = {
    '浙江-嘉兴': '事业一部',
    '浙江-湖州': '事业一部',
    '浙江-金华': '事业一部',
    '浙江-丽水': '事业一部',
    '浙江-衢州': '事业一部',
    '安徽': '事业一部',
    '湖北': '事业一部',
    '广东': '事业一部',
    '广西': '事业一部',
    '海南': '事业一部',
    '福建': '事业一部',
    '浙江-温州': '事业二部',
    '浙江-台州': '事业二部',
    '山东': '事业二部',
    '重庆': '事业二部',
    '四川': '事业二部',
    '贵州': '事业二部',
    '云南': '事业二部',
    '浙江-绍兴': '事业三部',
    '浙江-宁波': '事业三部',
    '浙江-舟山': '事业三部',
    '江苏': '事业三部',
    '河北': '事业三部',
    '北京': '事业三部',
    '黑龙江': '事业三部',
    '吉林': '事业三部',
    '辽宁': '事业三部',
    '天津': '事业三部',
    '陕西': '事业三部',
    '山西': '事业三部',
    '新疆': '事业三部',
    '青海': '事业三部',
    '甘肃': '事业三部',
    '宁夏': '事业三部',
    '河南': '事业三部',
    '内蒙古': '事业三部',
    '浙江-杭州': '事业四部',
    '上海': '事业四部',
    '湖南': '事业四部',
    '江西': '事业四部',
}


# 处理区域深耕TOP级客户表头
def handelTopHead(sheet: Worksheet, thisMonday: datetime.date, thisSunday: datetime.date, startRow: int = 4):
    """
    处理区域深耕TOP级客户表头
    :param sheet: 当前工作表对象
    :param thisMonday: 本周周一对应的日期
    :param thisSunday: 本周周日对应的日期
    :param startRow: 数据开始行
    :return:
    """
    # 插入表头
    oneHead = datetime_utils.formatDate(thisMonday) + '-' + datetime_utils.formatDate(thisSunday) + ' 区域深耕TOP级客户'
    xlsx_utils.insertContent(sheet=sheet, row=1, col=1, content=oneHead, rowHeight=30, fontSize=15, isBold=True)
    xlsx_utils.insertContent(sheet=sheet, row=2, col=1, content='区域', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=1, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=2, content='上周信息总量', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=2, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=3, content='目前信息总量', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=3, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=4, content='本周拜访计划', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=4, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=5, content='执行情况', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=5, content='可跟进', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=6, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=6, content='关闭', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=7, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=7, content='转为锁定', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=8, content='下周预计拜访量', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=8, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    contentList = ['杭州', '湖州', '嘉兴', '金华', '丽水', '宁波', '衢州', '绍兴', '台州', '温州', '合计']
    for content in contentList:
        xlsx_utils.insertContent(sheet=sheet, row=startRow, col=1, content=content, rowHeight=30, fgColor='00D7D7D7',
                                 bgColor='00D7D7D7')
        for col in range(2, 9):
            xlsx_utils.insertContent(sheet=sheet, row=startRow, col=col, rowHeight=30, fgColor='00D7D7D7',
                                     bgColor='00D7D7D7')
        startRow += 1
    # 合并单元格
    sheet.merge_cells(start_column=1, end_column=8, start_row=1, end_row=1)
    sheet.merge_cells(start_column=1, end_column=1, start_row=2, end_row=3)
    sheet.merge_cells(start_column=2, end_column=2, start_row=2, end_row=3)
    sheet.merge_cells(start_column=3, end_column=3, start_row=2, end_row=3)
    sheet.merge_cells(start_column=4, end_column=4, start_row=2, end_row=3)
    sheet.merge_cells(start_column=5, end_column=7, start_row=2, end_row=2)
    sheet.merge_cells(start_column=8, end_column=8, start_row=2, end_row=3)


# 处理蓝绿体系表头
def handelLanLvHead(sheet: Worksheet, thisMonday: datetime.date, thisSunday: datetime.date, startRow: int = 4):
    """
    处理蓝绿体系表头
    :param sheet: 当前工作表对象
    :param thisMonday: 本周周一对应的日期
    :param thisSunday: 本周周日对应的日期
    :param startRow: 数据开始行
    :return:
    """
    # 插入表头
    oneHead = datetime_utils.formatDate(thisMonday) + '-' + datetime_utils.formatDate(thisSunday) + ' 蓝绿体系'
    xlsx_utils.insertContent(sheet=sheet, row=1, col=1, content=oneHead, rowHeight=30, fontSize=15, isBold=True)
    xlsx_utils.insertContent(sheet=sheet, row=2, col=1, content='部门', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=1, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=2, content='目前洽谈总量', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=2, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=3, content='本周拜访计划', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=3, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=4, content='执行情况', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=4, content='可跟进', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=5, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=5, content='关闭', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=6, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=6, content='转为锁定', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=7, content='下周预计拜访量', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=7, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    contentList = ['事业一部', '事业二部', '事业三部', '事业四部', '合计']
    for content in contentList:
        xlsx_utils.insertContent(sheet=sheet, row=startRow, col=1, content=content, rowHeight=30, fgColor='00D7D7D7',
                                 bgColor='00D7D7D7')
        for col in range(2, 8):
            xlsx_utils.insertContent(sheet=sheet, row=startRow, col=col, rowHeight=30, fgColor='00D7D7D7',
                                     bgColor='00D7D7D7')
        startRow += 1
    # 合并单元格
    sheet.merge_cells(start_column=1, end_column=7, start_row=1, end_row=1)
    sheet.merge_cells(start_column=1, end_column=1, start_row=2, end_row=3)
    sheet.merge_cells(start_column=2, end_column=2, start_row=2, end_row=3)
    sheet.merge_cells(start_column=3, end_column=3, start_row=2, end_row=3)
    sheet.merge_cells(start_column=4, end_column=6, start_row=2, end_row=2)
    sheet.merge_cells(start_column=7, end_column=7, start_row=2, end_row=3)


# 处理绿城物业全委新签锁定项目表头
def handelQuanWeiXinQianHead(sheet: Worksheet, thisMonday: datetime.date, thisSunday: datetime.date, startRow: int = 4):
    """
    处理绿城物业全委新签锁定项目表头
    :param sheet: 当前工作表对象
    :param thisMonday: 本周周一对应的日期
    :param thisSunday: 本周周日对应的日期
    :param startRow: 数据开始行
    :return:
    """
    # 插入表头
    oneHead = datetime_utils.formatDate(thisMonday) + '-' + datetime_utils.formatDate(thisSunday) + ' 绿城物业全委新签锁定项目'
    xlsx_utils.insertContent(sheet=sheet, row=1, col=1, content=oneHead, rowHeight=30, fontSize=15, isBold=True)
    xlsx_utils.insertContent(sheet=sheet, row=2, col=1, content='部门', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=1, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=2, content='本周拜访计划', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=2, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=3, content='执行情况', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=3, content='可跟进', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=4, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=4, content='关闭', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=5, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=5, content='转为锁定', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=2, col=6, content='下周预计拜访量', rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=3, col=6, rowHeight=30, fgColor='00D7D7D7', bgColor='00D7D7D7')
    contentList = ['事业一部', '事业二部', '事业三部', '事业四部', '合计']
    for content in contentList:
        xlsx_utils.insertContent(sheet=sheet, row=startRow, col=1, content=content, rowHeight=30, fgColor='00D7D7D7',
                                 bgColor='00D7D7D7')
        for col in range(2, 7):
            xlsx_utils.insertContent(sheet=sheet, row=startRow, col=col, rowHeight=30, fgColor='00D7D7D7',
                                     bgColor='00D7D7D7')
        startRow += 1
    # 合并单元格
    sheet.merge_cells(start_column=1, end_column=6, start_row=1, end_row=1)
    sheet.merge_cells(start_column=1, end_column=1, start_row=2, end_row=3)
    sheet.merge_cells(start_column=2, end_column=2, start_row=2, end_row=3)
    sheet.merge_cells(start_column=3, end_column=5, start_row=2, end_row=2)
    sheet.merge_cells(start_column=6, end_column=6, start_row=2, end_row=3)


# 处理输出文件
def handelOutPutFile(fileName: str):
    """
    处理输出文件
    :param fileName: 源文件（包含路径）
    :return: 输出文件（包含路径）
    """
    # 处理输出文件的名称（包含路径）
    path, _ = os.path.split(fileName)
    thisMonday, thisSunday = datetime_utils.getThisWeek()
    thisMondayStr = datetime_utils.delDateYear(thisMonday, '月', '日')
    thisSundayStr = datetime_utils.delDateYear(thisSunday, '月', '日')
    outFileName = os.path.join(path, '市场拓展周报数据(' + thisMondayStr + '-' + thisSundayStr + ').xlsx')
    # 判断输出文件是否存在
    isExists = os.path.exists(outFileName)
    if not isExists:
        wb = Workbook()
        # 处理区域深耕TOP级客户表头
        wb[wb.sheetnames[0]].title = '区域深耕TOP级客户'
        handelTopHead(sheet=wb['区域深耕TOP级客户'], thisMonday=thisMonday, thisSunday=thisSunday)
        # 处理蓝绿体系表头
        wb.create_sheet('蓝绿体系')
        handelLanLvHead(sheet=wb['蓝绿体系'], thisMonday=thisMonday, thisSunday=thisSunday)
        # 处理绿城物业全委新签锁定项目表头
        wb.create_sheet('绿城物业全委新签锁定项目')
        handelQuanWeiXinQianHead(sheet=wb['绿城物业全委新签锁定项目'], thisMonday=thisMonday, thisSunday=thisSunday)
        # 保存输出文件
        wb.save(outFileName)
    return outFileName


# 更新表单数据
def updateFormData(sheet: Worksheet, destDict: dict, col: int, isDistinct: bool = True, startRow: int = 4):
    """
    更新表单数据
    :param sheet: 当前工作表对象
    :param destDict: 字典
    :param col: 列
    :param isDistinct: 是否去掉重复数据
    :param startRow: 数据开始行
    :return:
    """
    total = 0
    colList = xlsx_utils.getColList(sheet=sheet, row=startRow, col=1)
    totalRow = colList.index('合计') + startRow
    # 循环遍历，写入
    for key, values in destDict.items():
        row = colList.index(key) + startRow
        if isDistinct:
            count = len(set(values))
        else:
            count = len(values)
        total += count
        xlsx_utils.insertContent(sheet=sheet, row=row, col=col, content=count, rowHeight=30, fgColor='00D7D7D7',
                                 bgColor='00D7D7D7')
    xlsx_utils.insertContent(sheet=sheet, row=totalRow, col=col, content=total, rowHeight=30, fgColor='00D7D7D7',
                             bgColor='00D7D7D7')


# 处理旧文件
def handelOldFiles():
    fileName = tkinter.filedialog.askopenfilename(filetypes=[('xlsx files', '*.xlsx')])
    if fileName:
        # 选择了文件
        tkinter_utils.clearText(text=text, window=window)
        # 展示已选文件
        tkinter_utils.updateText(content='上周拓展信息表：' + fileName, text=text, window=window)
        # 处理输出文件
        outFileName = handelOutPutFile(fileName)
        # 打开输出文件
        outWb = load_workbook(outFileName)
        # 打开源xlsx表
        wb = load_workbook(fileName)
        # 已存在的全部工作簿
        sheetNames = wb.sheetnames
        # 处理指定的工作簿
        sheet_name = sheetNames[0]
        sheet = wb[sheet_name]
        # 城市和城市对应的公司列表
        companyDict = {}
        # 循环遍历，读取，判断
        startRow = 5
        for row in range(startRow, sheet.max_row + 1):
            # 城市
            city = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['B' + str(row)].column).value
            # 来源
            source = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['F' + str(row)].column).value
            # 公司名称
            company = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['H' + str(row)].column).value
            if source == '深耕top':
                if company is not None:
                    if city in companyDict:
                        companyList = companyDict[city]
                        companyList.append(company)
                        companyDict[city] = companyList
                    else:
                        companyDict[city] = [company]
        # 处理区域深耕TOP级客户表单数据
        updateFormData(sheet=outWb['区域深耕TOP级客户'], destDict=companyDict, col=2)
        # 保存输出文件
        outWb.save(outFileName)
        # 关闭文件
        wb.close()
        outWb.close()
        newBtn["state"] = NORMAL
    else:
        # 取消了选择
        messagebox.showinfo('提示', '未选择上周拓展信息表！！！')


# 处理新文件
def handelNewFiles():
    fileName = tkinter.filedialog.askopenfilename(filetypes=[('xlsx files', '*.xlsx')])
    if fileName:
        # 选择了文件
        # 展示已选文件
        tkinter_utils.updateText(content='本周拓展信息表：' + fileName, text=text, window=window)
        # 处理输出文件
        outFileName = handelOutPutFile(fileName)
        # 打开输出文件
        outWb = load_workbook(outFileName)
        # 打开源xlsx表
        wb = load_workbook(fileName)
        # 已存在的全部工作簿
        sheetNames = wb.sheetnames
        # 处理指定的工作簿
        sheet_name = sheetNames[0]
        sheet = wb[sheet_name]
        # 城市和城市对应的公司列表
        topCompanyDict = {}
        lanlvCompanyDict = {}
        # 城市和城市对应的本周拜访计划列表
        topThisPlanDict = {}
        lanlvThisPlanDict = {}
        # 城市和城市对应的本周可跟进
        topThisFollowDict = {}
        lanlvThisFollowDict = {}
        # 城市和城市对应的本周关闭
        topThisCloseDict = {}
        lanlvThisCloseDict = {}
        # 城市和城市对应的本周转为锁定
        topThisLockDict = {}
        lanlvThisLockDict = {}
        # 城市和城市对应的下周拜访计划
        topNextPlanDict = {}
        lanlvNextPlanDict = {}
        # 查找本周拜访计划、本周执行情况、下周拜访计划所在的列名
        startRow = 5
        rowList = xlsx_utils.getRowList(sheet=sheet, row=startRow - 1)
        thisMonday, thisSunday = datetime_utils.getThisWeek()
        nextMonday, nextSunday = datetime_utils.getNextWeek()
        thisPlanColLetter = ''
        thisExecuteColLetter = ''
        nextPlanColLetter = ''
        for rowItem in rowList:
            thisWeek = datetime_utils.delDateYear(thisMonday, '.') + '-' + datetime_utils.delDateYear(thisSunday, '.')
            nextWeek = datetime_utils.delDateYear(nextMonday, '.') + '-' + datetime_utils.delDateYear(nextSunday, '.')
            if thisWeek in rowItem:
                thisIndex = rowList.index(rowItem)
                thisPlanColLetter = sheet.cell(startRow - 1, thisIndex + 1).column_letter
                thisExecuteColLetter = sheet.cell(startRow - 1, thisIndex + 3).column_letter
            if nextWeek in rowItem:
                nextIndex = rowList.index(rowItem)
                nextPlanColLetter = sheet.cell(startRow - 1, nextIndex + 1).column_letter
        # 循环遍历，读取，判断
        for row in range(startRow, sheet.max_row + 1):
            # 省会
            province = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['A' + str(row)].column).value
            # 城市
            city = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['B' + str(row)].column).value
            # 来源
            source = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['F' + str(row)].column).value
            # 公司名称
            lanlvCompany = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['G' + str(row)].column).value
            topCompany = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['H' + str(row)].column).value
            # 本周跟进计划
            thisPlan = xlsx_utils.parserMergedCell(sheet=sheet, row=row,
                                                   col=sheet[thisPlanColLetter + str(row)].column).value
            # 本周执行情况
            thisExecute = xlsx_utils.parserMergedCell(sheet=sheet, row=row,
                                                      col=sheet[thisExecuteColLetter + str(row)].column).value
            # 下周跟进计划
            nextPlan = xlsx_utils.parserMergedCell(sheet=sheet, row=row,
                                                   col=sheet[nextPlanColLetter + str(row)].column).value
            if source == '深耕top':
                # 处理公司信息
                if topCompany is not None:
                    if city in topCompanyDict:
                        companyList = topCompanyDict[city]
                        companyList.append(topCompany)
                        topCompanyDict[city] = companyList
                    else:
                        topCompanyDict[city] = [topCompany]
                # 处理本周计划
                if thisPlan is not None:
                    if city in topThisPlanDict:
                        thisPlanList = topThisPlanDict[city]
                        thisPlanList.append(thisPlan)
                        topThisPlanDict[city] = thisPlanList
                    else:
                        topThisPlanDict[city] = [thisPlan]
                # 处理本周执行情况
                if thisExecute == '可跟进':
                    if city in topThisFollowDict:
                        thisFollowList = topThisFollowDict[city]
                        thisFollowList.append(thisExecute)
                        topThisFollowDict[city] = thisFollowList
                    else:
                        topThisFollowDict[city] = [thisExecute]
                elif thisExecute == '关闭':
                    if city in topThisCloseDict:
                        thisCloseList = topThisCloseDict[city]
                        thisCloseList.append(thisExecute)
                        topThisCloseDict[city] = thisCloseList
                    else:
                        topThisCloseDict[city] = [thisExecute]
                elif thisExecute == '锁定':
                    if city in topThisLockDict:
                        thisLockList = topThisLockDict[city]
                        thisLockList.append(thisExecute)
                        topThisLockDict[city] = thisLockList
                    else:
                        topThisLockDict[city] = [thisExecute]
                # 处理下周计划
                if nextPlan is not None:
                    if city in topNextPlanDict:
                        nextPlanList = topNextPlanDict[city]
                        nextPlanList.append(nextPlan)
                        topNextPlanDict[city] = nextPlanList
                    else:
                        topNextPlanDict[city] = [nextPlan]
            elif source == '蓝城' or source == '绿城小镇' or source == '绿管' or source == '绿中':
                if province == '浙江':
                    pcKey = province + '-' + city
                    if pcKey in provinceCityDict:
                        bmKey = provinceCityDict[pcKey]
                    else:
                        messagebox.showerror('错误', '找不到地区对应的部门数据，请检查配置：' + pcKey + '\n')
                        return
                else:
                    if province in provinceCityDict:
                        bmKey = provinceCityDict[province]
                    else:
                        messagebox.showerror('错误', '找不到地区对应的部门数据，请检查配置：' + province + '\n')
                        return
                # 处理公司信息
                if bmKey in lanlvCompanyDict:
                    companyList = lanlvCompanyDict[bmKey]
                    companyList.append(lanlvCompany)
                    lanlvCompanyDict[bmKey] = companyList
                else:
                    lanlvCompanyDict[bmKey] = [lanlvCompany]
                # 处理本周计划
                if thisPlan is not None:
                    if bmKey in lanlvThisPlanDict:
                        thisPlanList = lanlvThisPlanDict[bmKey]
                        thisPlanList.append(thisPlan)
                        lanlvThisPlanDict[bmKey] = thisPlanList
                    else:
                        lanlvThisPlanDict[bmKey] = [thisPlan]
                # 处理本周执行情况
                if thisExecute == '可跟进':
                    if bmKey in lanlvThisFollowDict:
                        thisFollowList = lanlvThisFollowDict[bmKey]
                        thisFollowList.append(thisExecute)
                        lanlvThisFollowDict[bmKey] = thisFollowList
                    else:
                        lanlvThisFollowDict[bmKey] = [thisExecute]
                elif thisExecute == '关闭':
                    if bmKey in lanlvThisCloseDict:
                        thisCloseList = lanlvThisCloseDict[bmKey]
                        thisCloseList.append(thisExecute)
                        lanlvThisCloseDict[bmKey] = thisCloseList
                    else:
                        lanlvThisCloseDict[bmKey] = [thisExecute]
                elif thisExecute == '锁定':
                    if bmKey in lanlvThisLockDict:
                        thisLockList = lanlvThisLockDict[bmKey]
                        thisLockList.append(thisExecute)
                        lanlvThisLockDict[bmKey] = thisLockList
                    else:
                        lanlvThisLockDict[bmKey] = [thisExecute]
                    # 处理下周计划
                if nextPlan is not None:
                    if bmKey in lanlvNextPlanDict:
                        nextPlanList = lanlvNextPlanDict[bmKey]
                        nextPlanList.append(nextPlan)
                        lanlvNextPlanDict[bmKey] = nextPlanList
                    else:
                        lanlvNextPlanDict[bmKey] = [nextPlan]
        # 处理区域深耕TOP级客户表单数据
        updateFormData(sheet=outWb['区域深耕TOP级客户'], destDict=topCompanyDict, col=3)
        updateFormData(sheet=outWb['区域深耕TOP级客户'], destDict=topThisPlanDict, col=4, isDistinct=False)
        updateFormData(sheet=outWb['区域深耕TOP级客户'], destDict=topThisFollowDict, col=5, isDistinct=False)
        updateFormData(sheet=outWb['区域深耕TOP级客户'], destDict=topThisCloseDict, col=6, isDistinct=False)
        updateFormData(sheet=outWb['区域深耕TOP级客户'], destDict=topThisLockDict, col=7, isDistinct=False)
        updateFormData(sheet=outWb['区域深耕TOP级客户'], destDict=topNextPlanDict, col=8, isDistinct=False)
        # 处理蓝绿体系表单数据
        updateFormData(sheet=outWb['蓝绿体系'], destDict=lanlvCompanyDict, col=2, isDistinct=False)
        updateFormData(sheet=outWb['蓝绿体系'], destDict=lanlvThisPlanDict, col=3, isDistinct=False)
        updateFormData(sheet=outWb['蓝绿体系'], destDict=lanlvThisFollowDict, col=4, isDistinct=False)
        updateFormData(sheet=outWb['蓝绿体系'], destDict=lanlvThisCloseDict, col=5, isDistinct=False)
        updateFormData(sheet=outWb['蓝绿体系'], destDict=lanlvThisLockDict, col=6, isDistinct=False)
        updateFormData(sheet=outWb['蓝绿体系'], destDict=lanlvNextPlanDict, col=7, isDistinct=False)
        # 保存输出文件
        outWb.save(outFileName)
        # 关闭文件
        wb.close()
        outWb.close()
        newBtn["state"] = DISABLED
        messagebox.showinfo('提示', '文件处理完成！！！')
    else:
        # 取消了选择
        messagebox.showinfo('提示', '未选择本周拓展信息表！！！')


oldBtn = tkinter.Button(window, text='选择上周拓展信息表', command=handelOldFiles)
oldBtn.place(x=200, y=10)
newBtn = tkinter.Button(window, text='选择本周拓展信息表', command=handelNewFiles)
newBtn.place(x=450, y=10)
newBtn["state"] = DISABLED
text = tkinter.Text(window)
text.place(x=120, y=60)
text.config(state=DISABLED)
window.mainloop()
