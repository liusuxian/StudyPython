from requests_html import HTMLSession
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

url = "https://www.cwl.gov.cn/ygkj/wqkjgg/ssq/"
session = HTMLSession()


def fetch_ssq_data(max_pages=1):
    data = []  # 用于存储所有数据的列表
    r = session.get(url)
    r.html.render(scrolldown=5, sleep=1)

    page = 1
    while page <= max_pages:
        items = r.html.xpath("//table/tbody/tr")

        # 如果没有更多数据，跳出循环
        if not items:
            break

        print(f"Page {page}: {len(items)} items found.")

        for item in items:
            # 在每次循环中，找到所需信息并打印
            tdList = item.find("td", first=False)
            period_number = tdList[0].text
            date = tdList[1].text
            first_prize = tdList[3].text
            single_award = tdList[4].text
            second_prize = tdList[5].text
            second_single_award = tdList[6].text
            winning_numbers = item.find(".qiu", first=True).text.replace("\n", " ")

            # 将数据添加到列表中
            data.append(
                {
                    "期号": int(period_number),
                    "开奖日期": date,
                    "开奖号码": winning_numbers,
                    "一等奖注数": int(first_prize.replace(",", "")),
                    "一等奖单注奖金": int(single_award.replace(",", "")),
                    "二等奖注数": int(second_prize.replace(",", "")),
                    "二等奖单注奖金": int(second_single_award.replace(",", "")),
                }
            )

        # 使用 JavaScript 模拟点击“下一页”按钮
        next_button = r.html.find("a.layui-laypage-next", first=True)
        if next_button:
            next_page_script = (
                f"document.querySelector('a.layui-laypage-next').click();"
            )
            r.html.render(script=next_page_script, scrolldown=5, sleep=1)
            page += 1
        else:
            break

    # 将数据写入 Excel
    df = pd.DataFrame(data)
    excel_file = "shuangseqiu_data.xlsx"
    df.to_excel(excel_file, index=False)

    # 加载工作簿并获取工作表
    wb = load_workbook(excel_file)
    ws = wb.active

    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 设置数据行框线和居中对齐
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # 自适应列宽，考虑所有单元格和表头
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                # 计算每个单元格的长度，中文字符按2倍长度计算
                cell_length = sum(
                    2 if ord(char) > 127 else 1 for char in str(cell.value)
                )
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        # 增加宽度以确保内容完全显示
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 保存工作簿
    wb.save(excel_file)
    print("数据已写入并格式化到 shuangseqiu_data.xlsx")


if __name__ == "__main__":
    fetch_ssq_data(max_pages=5)
