# coding=utf-8
from tkinter import messagebox
from openpyxl import load_workbook
import tkinter.filedialog

from StudyPython.src import tkinter_utils, xlsx_utils

window = tkinter_utils.createWindow('全委子表处理工具')
# 省会和城市所属的事业部
provinceCityDict = {
    '浙江-嘉兴': '战区四部',
    '浙江-湖州': '战区四部',
    '浙江-金华': '战区三部',
    '浙江-丽水': '战区三部',
    '浙江-衢州': '战区三部',
    '安徽': '战区四部',
    '浙江-温州': '战区二部',
    '浙江-台州': '战区五部',
    '山东': '战区五部',
    '重庆': '战区五部',
    '四川': '战区五部',
    '浙江-绍兴': '战区六部',
    '浙江-宁波': '战区一部',
    '浙江-舟山': '战区一部',
    '江苏': '战区六部',
    '河北': '战区六部',
    '北京': '战区六部',
    '天津': '战区六部',
    '浙江-杭州（外域）': '战区二部',
    '浙江-杭州（市内）': '战区七部',
    '湖南': '战区七部',
    '江西': '战区七部',
}


def handelFiles():
    fileName = tkinter.filedialog.askopenfilename(filetypes=[('xlsx files', '*.xlsx')])
    if fileName:
        # 选择了文件
        tkinter_utils.clearText(text=text, window=window)
        # 展示已选文件
        tkinter_utils.updateText(content='已选文件：' + fileName, text=text, window=window)
        # 打开源xlsx表
        wb = load_workbook(fileName)
        # 已存在的全部工作簿
        sheetNames = wb.sheetnames
        # 处理指定的工作簿
        if '全委新签表' not in sheetNames:
            errStr = '文件：' + fileName + ' 没有需要处理的工作簿或工作簿名称不正确' + '\n'
            messagebox.showerror('错误', errStr)
            return
        sheet_name = '全委新签表'
        sheet = wb[sheet_name]
        # 循环遍历，读取，判断
        height = sheet.row_dimensions[sheet.max_column].height
        targetColumn = sheet.max_column + 1
        xlsx_utils.insertContent(sheet=sheet, row=1, col=targetColumn, content='所属事业部', rowHeight=height)
        # 合并单元格
        sheet.merge_cells(start_column=targetColumn, end_column=targetColumn, start_row=2, end_row=4)
        startRow = 5
        for row in range(startRow, sheet.max_row + 1):
            # 省会
            province = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['E' + str(row)].column).value
            # 城市
            city = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['F' + str(row)].column).value
            # 项目名称
            project = xlsx_utils.parserMergedCell(sheet=sheet, row=row, col=sheet['C' + str(row)].column).value
            if province == '浙江':
                pcKey = str(province) + '-' + str(city)
                if pcKey in provinceCityDict:
                    bmKey = provinceCityDict[pcKey]
                    height = sheet.row_dimensions[targetColumn - 1].height
                    xlsx_utils.insertContent(sheet=sheet, row=row, col=targetColumn, content=bmKey, rowHeight=height)
                else:
                    tkinter_utils.updateText(content='找不到地区对应的部门数据，请检查配置：' + str(pcKey) + ' ' + str(project), text=text,
                                             window=window)
            else:
                if province in provinceCityDict:
                    bmKey = provinceCityDict[province]
                    height = sheet.row_dimensions[targetColumn - 1].height
                    xlsx_utils.insertContent(sheet=sheet, row=row, col=targetColumn, content=bmKey, rowHeight=height)
                else:
                    tkinter_utils.updateText(content='找不到地区对应的部门数据，请检查配置：' + str(province) + ' ' + str(project),
                                             text=text, window=window)
        # 保存输出文件
        wb.save(fileName)
        # 关闭文件
        wb.close()
        messagebox.showinfo('提示', '文件处理完成！！！')
    else:
        # 取消了选择
        messagebox.showinfo('提示', '未选择文件！！！')


btn = tkinter.Button(window, text='选择文件', command=handelFiles)
btn.pack()
text = tkinter.Text(window)
text.pack()
window.mainloop()
