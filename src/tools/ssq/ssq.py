from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from collections import Counter
from bs4 import BeautifulSoup
import time

chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.binary_location = "chromedriver"
driver = webdriver.Chrome(options=chrome_options)
url = "https://www.cwl.gov.cn/ygkj/wqkjgg/ssq/"


def fetch_ssq_data(max_pages=1):
    data = []  # 用于存储所有数据的列表
    driver.get(url)

    red_counter = Counter()
    blue_counter = Counter()

    page = 1
    while page <= max_pages:
        soup = BeautifulSoup(driver.page_source, "html.parser")
        items = soup.select("table tbody tr")

        # 如果没有更多数据，跳出循环
        if not items:
            break

        print(f"Page {page}: {len(items)} items found.")

        for item in items:
            # 在每次循环中，找到所需信息并打印
            tdList = item.find_all("td")
            period_number = tdList[0].text
            date = tdList[1].text
            first_prize = tdList[3].text if tdList[3].text != "_" else "0"
            single_award = tdList[4].text if tdList[4].text != "_" else "0"
            second_prize = tdList[5].text if tdList[5].text != "_" else "0"
            second_single_award = tdList[6].text if tdList[6].text != "_" else "0"
            numbers_div = tdList[2].find("div", class_="qiu")
            numbers_divs = numbers_div.find_all("div")
            winning_numbers = " ".join([num_div.text for num_div in numbers_divs])

            # 分解红球和篮球
            numbers = winning_numbers.split()
            red_balls = numbers[:6]
            blue_ball = numbers[6]

            # 更新红球和篮球的统计
            red_counter.update(red_balls)
            blue_counter.update([blue_ball])

            # 将数据添加到列表中
            data.append(
                {
                    "期号": int(period_number),
                    "开奖日期": date,
                    "开奖号码": winning_numbers,
                    "一等奖注数": int(first_prize.replace(",", "")),
                    "一等奖单注奖金": int(single_award.replace(",", "")),
                    "二等奖注数": int(second_prize.replace(",", "")),
                    "二等奖单注奖金": int(second_single_award.replace(",", "")),
                }
            )

        # 使用 JavaScript 模拟点击“下一页”按钮
        try:
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a.layui-laypage-next"))
            )
            if next_button:
                print(f"点击进入第 {page + 1} 页")
                next_button.click()
                time.sleep(1)
                page += 1
            else:
                print("未找到下一页按钮，结束循环")
                break
        except Exception as e:
            print("错误信息：", str(e))
            break

    # 将数据写入 Excel
    df = pd.DataFrame(data)
    excel_file = "shuangseqiu_data.xlsx"
    df.to_excel(excel_file, index=False)

    # 加载工作簿并获取工作表
    wb = load_workbook(excel_file)
    ws = wb.active

    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 设置数据行框线和居中对齐
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # 自适应列宽，考虑所有单元格和表头
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                # 计算每个单元格的长度，中文字符按2倍长度计算
                cell_length = sum(
                    2 if ord(char) > 127 else 1 for char in str(cell.value)
                )
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        # 增加宽度以确保内容完全显示
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 保存工作簿
    wb.save(excel_file)
    print("数据已写入并格式化到 shuangseqiu_data.xlsx")
    return red_counter, blue_counter


def predict_numbers(red_counter, blue_counter):
    # 获取出现次数最多的红球和篮球
    most_common_red = [num for num, _ in red_counter.most_common(7)]
    most_common_blue = [num for num, _ in blue_counter.most_common(2)]
    return most_common_red, most_common_blue


if __name__ == "__main__":
    red_counter, blue_counter = fetch_ssq_data(max_pages=5)
    predicted_red, predicted_blue = predict_numbers(red_counter, blue_counter)
    print(f"预测红球: {sorted(predicted_red)}, 预测蓝球: {sorted(predicted_blue)}")
