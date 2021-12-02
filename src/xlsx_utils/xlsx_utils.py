from openpyxl.styles import Font, Color, Alignment, Side, Border, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell


# 检查是否为合并单元格并获取对应行列单元格的值
# 如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值，否则直接返回该单元格的值
def parserMergedCell(sheet: Worksheet, row: int, col: int):
    """
    检查是否为合并单元格并获取对应行列单元格的值
    如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值，否则直接返回该单元格的值
    :param sheet: 当前工作表对象
    :param row: 需要获取的单元格所在行
    :param col: 需要获取的单元格所在列
    :return:
    """
    cell = sheet.cell(row=row, column=col)
    # 判断该单元格是否为合并单元格
    if isinstance(cell, MergedCell):
        # 循环查找该单元格所属的合并区域
        for merged_range in sheet.merged_cell_ranges:
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell


# 获取工作表从某一列开始指定的行的内容
def getRowList(sheet: Worksheet, row: int = 1, col: int = 1):
    """
    获取工作表从某一列开始指定的行的内容
    :param sheet: 当前工作表对象
    :param row: 行
    :param col: 列
    :return: 指定的行的内容
    """
    rowList = []
    for i in range(col, sheet.max_column + 1):
        val = parserMergedCell(sheet=sheet, row=row, col=i).value
        rowList.append(val)
    return rowList


# 获取工作表从某一行开始指定的列的内容
def getColList(sheet: Worksheet, row: int = 1, col: int = 1):
    """
    获取工作表从某一行开始指定的列的内容
    :param sheet: 当前工作表对象
    :param row: 行
    :param col: 列
    :return: 指定的列的内容
    """
    colList = []
    for i in range(row, sheet.max_row + 1):
        val = parserMergedCell(sheet=sheet, row=i, col=col).value
        colList.append(val)
    return colList


# 在指定单元格内插入内容和格式
def insertContent(
        sheet: Worksheet,
        row: int = 1,
        col: int = 1,
        content: str | int = None,
        colWidth: int = 15,
        rowHeight: int = 10,
        fontName: str = '宋体',
        fontSize: int = 10,
        fontColor: str = '00000000',
        isBold: bool = False,
        isItalic: bool = False,
        fgColor: str = '00d9d2e9',
        bgColor: str = '00d9d2e9',
        sideColor: str = '00000000'
):
    """
    在指定单元格内插入内容和格式
    :param sheet: 当前工作表对象
    :param row: 行
    :param col: 列
    :param content: 单元格内容
    :param colWidth: 列宽
    :param rowHeight: 行高
    :param fontName: 字体
    :param fontSize: 字体大小
    :param fontColor: 字体颜色
    :param isBold: 是否粗体
    :param isItalic: 是否斜体
    :param fgColor: 前景色
    :param bgColor: 后景色
    :param sideColor: 边框颜色
    :return:
    """
    cell = sheet.cell(row=row, column=col)
    coord = cell.coordinate
    # 设置单元格内容
    cell.value = content
    # 设置单元格宽度
    sheet.column_dimensions[get_column_letter(col)].width = colWidth
    # 设置单元格高度
    sheet.row_dimensions[row].height = rowHeight
    # 设置单元格字体
    sheet[coord].font = Font(name=fontName, size=fontSize, color=Color(rgb=fontColor), b=isBold, i=isItalic)
    # 设置单元格对齐（水平居中，垂直居中）
    sheet[coord].alignment = Alignment(horizontal='center', vertical='center')
    # 设置边框
    side = Side(style='thin', color=Color(rgb=sideColor))
    sheet[coord].border = Border(left=side, right=side, top=side, bottom=side)
    # 设置单元格填充
    sheet[coord].fill = PatternFill(patternType='lightDown', fgColor=Color(rgb=fgColor), bgColor=Color(rgb=bgColor))
